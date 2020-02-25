# coding：<encoding name> ： # coding: utf-8

import os
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import colors

print("Which line of data would you like to input?")
# day = 3
day = int(input())

print("What is the answer key?")
print("Please type in UPPERCASE letters without space or comma.")
keys = ['D', 'B', 'D', 'B', 'D', 'D', 'D', 'B', 'B', 'B', 'B', 'C', 'C', 'A', 'C', 'D', 'B', 'B', 'D', 'B']
keys = input()
# DBDBDDDBBBBCCACDBBDB

awb = openpyxl.Workbook()
awsheet = awb.active

# first row
for i in range(4, 24):
    awsheet.cell(row = 1, column = i - 1, value = i - 3)
awsheet.cell(row = 1, column = 1).value = "学号"      # write in
awsheet.cell(row = 1, column = 2).value = "姓名"
awsheet.cell(row = 1, column = 23).value = "score"

# 赋字体对象
RedFont = Font(bold = True, color = colors.RED)

f = os.listdir(".")    # 获取当前目录下的所有文件名
length = len(f)
j = 2
for i in range(0, length):
    file_name = f[i]
    if os.path.splitext(file_name)[1] == '.xlsx' and os.path.splitext(file_name)[0] != 'aws':          # 文件后缀分离
        #[0]文件名，[1]后缀名
        # 如果是xlsx文件
        # print(file_name)
        score = 30
        awsheet.cell(row = j, column = 23, value = score)

        wb = openpyxl.load_workbook(file_name)       # workbook
        ws = wb.active      # worksheet
        # get name
        numName = ws.cell(row = day, column = 2).value
        name = ws.cell(row = day, column = 3).value
        
        awsheet.cell(row = j, column = 1).value = numName      # write in
        awsheet.cell(row = j, column = 2).value = name
        # print("学号： ", numName)
        # print("姓名： ", name)

        aws = []
        # get answer 1-20
        for i in range(4, 24):
            charaws = ws.cell(row = day, column = i).value.upper()      # turn into upper case
            aws.append(charaws) 
            awsheet.cell(row = j, column = i - 1, value = charaws)
            if charaws != keys[i - 4]:      # WA
                # thiscell.font.italic = True
                awsheet.cell(row = j, column = i - 1, value = charaws).font = RedFont
                if i<=14:
                    score = score - 1
                    awsheet.cell(row = j, column = 23, value = score)
                else:
                    score = score - 2                
                    awsheet.cell(row = j, column = 23, value = score)
        print (aws)     # 打印答案数组
        j = j + 1

# print(len(f))         # 目录下文件个数
# aws.cell(row = 1, column = 1).value = 'test'
awb.save('aws.xlsx')
print("Completed. Thanks for using.")